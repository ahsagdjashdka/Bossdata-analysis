import pymysql
from pymysql.cursors import DictCursor
import openpyxl
from openpyxl import load_workbook
import os
import csv


class DBManager:
    def __init__(self, host='localhost', user='root', password='123456',
                 database='job_data', charset='utf8mb4'):
        """初始化数据库连接"""
        self.connection = pymysql.connect(
            host=host,
            user=user,
            password=password,
            database=database,
            charset=charset,
            cursorclass=DictCursor
        )
        self.cursor = self.connection.cursor()

    def __enter__(self):
        """支持with语句"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """支持with语句"""
        self.close()

    def close(self):
        """关闭连接"""
        if self.cursor:
            self.cursor.close()
        if self.connection:
            self.connection.close()

    def execute(self, sql, args=None):
        """执行SQL语句"""
        try:
            self.cursor.execute(sql, args or ())
            self.connection.commit()
            return self.cursor
        except Exception as e:
            self.connection.rollback()
            raise e

    def create_database(self):
        """创建数据库"""
        try:
            # 临时连接到MySQL服务器创建数据库
            temp_conn = pymysql.connect(
                host='localhost',
                user='root',
                password='123456',
                charset='utf8mb4'
            )
            with temp_conn.cursor() as cursor:
                cursor.execute(
                    "CREATE DATABASE IF NOT EXISTS job_data CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
                print("数据库job_data创建成功或已存在")
            temp_conn.close()
        except Exception as e:
            print(f"创建数据库时出错: {str(e)}")
            raise

    def create_jobs_table(self):
        """创建职位信息表"""
        sql = """
        CREATE TABLE IF NOT EXISTS jobs (
            id INT AUTO_INCREMENT PRIMARY KEY,
            job_name VARCHAR(100) NOT NULL COMMENT '职位名称',
            company_add VARCHAR(200) COMMENT '公司地址',
            month VARCHAR(20) COMMENT '招聘月份',
            salary_min DECIMAL(10,2) COMMENT '最低薪资(千)',
            salary_max DECIMAL(10,2) COMMENT '最高薪资(千)',
            record VARCHAR(50) COMMENT '学历要求',
            company_name VARCHAR(100) COMMENT '公司名称',
            company_type VARCHAR(100) COMMENT '公司类型',
            scope VARCHAR(100) COMMENT '公司规模',
            tag VARCHAR(200) COMMENT '职位标签',
            welfare TEXT COMMENT '公司福利',
            create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间'
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='职位信息表';
        """
        self.execute(sql)
        print("数据表jobs创建成功或已存在")


def import_data_to_db(file_path):
    """将数据文件(Excel或CSV)导入数据库"""
    try:
        # 初始化数据库连接
        with DBManager() as db:
            # 创建数据库和表
            db.create_database()
            db.create_jobs_table()

            # 根据文件扩展名选择不同的读取方式
            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                # 处理Excel文件
                wb = load_workbook(file_path)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

            elif file_ext == '.csv':
                # 处理CSV文件
                with open(file_path, mode='r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  # 读取表头
                    rows = [row for row in reader]
            else:
                print(f"不支持的文件格式: {file_ext}")
                return False

            # 检查表头是否符合预期
            expected_headers = [
                'job_name', 'company_add', 'month', 'salary_min',
                'salary_max', 'record', 'company_name',
                'company_type', 'scope', 'tag', 'welfare'
            ]

            if headers != expected_headers:
                print(f"文件表头不匹配，期望: {expected_headers}，实际: {headers}")
                return False

            # 准备SQL插入语句
            sql = """
            INSERT INTO jobs (
                job_name, company_add, month, salary_min, salary_max,
                record, company_name, company_type, scope, tag, welfare
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            # 遍历数据行
            success_count = 0
            for row in rows:
                try:
                    # 处理数据（确保数据类型正确）
                    processed_row = list(row)

                    # 转换薪资为数字类型
                    processed_row[3] = float(processed_row[3]) if processed_row[3] else 0
                    processed_row[4] = float(processed_row[4]) if processed_row[4] else 0

                    # 执行插入
                    db.execute(sql, processed_row)
                    success_count += 1
                except Exception as e:
                    print(f"插入数据失败: {row}，错误: {str(e)}")
                    continue

            print(f"数据导入完成，成功插入{success_count}条记录")
            return True

    except Exception as e:
        print(f"导入数据时出错: {str(e)}")
        return False


if __name__ == "__main__":
    # 数据文件路径（支持.xlsx或.csv）
    data_file = "BossData_all.csv"  # 或 "job_data.xlsx"

    # 检查文件是否存在
    if not os.path.exists(data_file):
        print(f"错误: 文件 {data_file} 不存在")
    else:
        # 执行导入
        if import_data_to_db(data_file):
            print("数据导入成功！")
        else:
            print("数据导入失败！")